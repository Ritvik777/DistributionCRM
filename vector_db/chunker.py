from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from langchain_text_splitters import RecursiveCharacterTextSplitter

from config import CHUNK_SIZE, CHUNK_OVERLAP


@dataclass
class Chunk:
    """A piece of text plus structured metadata stored alongside it in Qdrant."""

    text: str
    metadata: dict = field(default_factory=dict)


_PRICE_HEADER_HINTS = re.compile(
    r"\b(product|item|sku|description|qty|quantity|unit|price|rate|amount|total|cost)\b",
    re.IGNORECASE,
)
_SKU_HINT = re.compile(r"\b(sku|item\s*code|item\s*no|part|code|model)\b", re.IGNORECASE)
_PRICE_HINT = re.compile(r"\b(price|rate|amount|total|cost|mrp)\b", re.IGNORECASE)
_NAME_HINT = re.compile(r"\b(product|description|name|service)\b", re.IGNORECASE)


def _splitter(chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> RecursiveCharacterTextSplitter:
    return RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
    )


def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    cleaned = text.strip()
    if not cleaned:
        return []
    if len(cleaned) <= chunk_size:
        return [cleaned]
    return _splitter(chunk_size, overlap).split_text(cleaned)


def _looks_numeric(value: Any) -> bool:
    text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def _to_number(value: Any) -> float:
    text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    return float(text)


def _normalize_header(cell: Any, index: int) -> str:
    if cell is None or not str(cell).strip():
        return f"Column {index + 1}"
    return str(cell).strip()


def _header_score(row: tuple[Any, ...] | list[Any]) -> int:
    values = [cell for cell in row if cell is not None and str(cell).strip()]
    if len(values) < 2:
        return 0

    joined = " ".join(str(value) for value in values)
    hint_bonus = 3 if _PRICE_HEADER_HINTS.search(joined) else 0
    text_cells = sum(1 for value in values if not _looks_numeric(value))
    return len(values) + text_cells + hint_bonus


def _find_header_index(rows: list[tuple[Any, ...]]) -> int:
    best_index = 0
    best_score = 0
    for index, row in enumerate(rows[:20]):
        score = _header_score(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _row_metadata(headers: list[str], row: tuple[Any, ...] | list[Any]) -> dict:
    """Pull useful fields (sku, product, price) out of a product row for filtering."""
    meta: dict = {}
    for header, cell in zip(headers, row):
        if cell is None or not str(cell).strip():
            continue
        value = str(cell).strip()
        if "sku" not in meta and _SKU_HINT.search(header):
            meta["sku"] = value
        if "product" not in meta and _NAME_HINT.search(header):
            meta["product"] = value
        if "price" not in meta and _PRICE_HINT.search(header) and _looks_numeric(value):
            meta["price"] = _to_number(value)
    return meta


def _format_product_chunk(sheet_label: str, headers: list[str], row: tuple[Any, ...] | list[Any]) -> str | None:
    pairs: list[str] = []
    for header, cell in zip(headers, row):
        if cell is None or not str(cell).strip():
            continue
        pairs.append(f"{header}: {cell}")

    if not pairs:
        return None

    return f"[{sheet_label}] " + " | ".join(pairs)


def _format_preamble_row(row: tuple[Any, ...] | list[Any]) -> str:
    cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
    if not cells:
        return ""
    if len(cells) >= 2 and len(cells) % 2 == 0:
        return " | ".join(f"{cells[index]}: {cells[index + 1]}" for index in range(0, len(cells), 2))
    return " ".join(cells)


def _chunks_from_tabular_rows(
    sheet_label: str,
    rows: list[list[Any]],
    base_metadata: dict | None = None,
) -> list[Chunk]:
    base_metadata = base_metadata or {}
    if not rows:
        return []

    tuples = [tuple(row) for row in rows]
    header_index = _find_header_index(tuples)
    header_row = tuples[header_index]
    headers = [_normalize_header(cell, index) for index, cell in enumerate(header_row)]

    chunks: list[Chunk] = []

    preamble = []
    for row in tuples[:header_index]:
        line = _format_preamble_row(row)
        if line:
            preamble.append(line)
    if preamble:
        text = f"[{sheet_label} — Quotation Info] " + " | ".join(preamble)
        chunks.append(Chunk(text, {**base_metadata, "type": "quotation_info", "sheet": sheet_label}))

    for row_number, row in enumerate(tuples[header_index + 1 :], start=1):
        text = _format_product_chunk(sheet_label, headers, row)
        if not text:
            continue
        meta = {**base_metadata, "type": "product", "sheet": sheet_label, "row": row_number}
        meta.update(_row_metadata(headers, row))
        for piece in (chunk_text(text) if len(text) > CHUNK_SIZE else [text]):
            chunks.append(Chunk(piece, dict(meta)))

    return chunks


def extract_chunks_from_pdf(pdf_file, source: str = "") -> list[Chunk]:
    """Extract tables as product rows and remaining text as prose, page by page."""
    import pdfplumber

    label = source or "PDF"
    chunks: list[Chunk] = []
    with pdfplumber.open(pdf_file) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            base = {"source": source, "page": page_number}
            tables = page.extract_tables() or []
            usable = False
            for table in tables:
                rows = [
                    [cell if cell is not None else "" for cell in row]
                    for row in table
                    if row and any((str(cell).strip() if cell is not None else "") for cell in row)
                ]
                if not rows:
                    continue
                usable = True
                chunks.extend(_chunks_from_tabular_rows(f"{label} p{page_number}", rows, base_metadata=dict(base)))

            if usable:
                continue

            text = (page.extract_text() or "").strip()
            for piece in chunk_text(text):
                chunks.append(Chunk(f"[Page {page_number}] {piece}", {**base, "type": "pdf"}))

    return chunks


def extract_chunks_from_excel(excel_file, source: str = "") -> list[Chunk]:
    from openpyxl import load_workbook

    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    chunks: list[Chunk] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            rows = [row for row in rows if any(cell is not None and str(cell).strip() for cell in row)]
            base = {"source": source}
            chunks.extend(_chunks_from_tabular_rows(worksheet.title, rows, base_metadata=base))
    finally:
        workbook.close()
    return chunks


def extract_chunks_from_csv(csv_file, source: str = "") -> list[Chunk]:
    content = csv_file.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return _chunks_from_tabular_rows("CSV", rows, base_metadata={"source": source})
